#!/usr/bin/env python3
"""
Enterprise Agent MCP Server — exposes BI, SW delivery, and data tools
to OpenClaw via stdio MCP transport so the Discord bot can invoke them.
"""
from __future__ import annotations

import json
import os
import sys
import textwrap
from pathlib import Path

# ensure project root is importable
sys.path.insert(0, str(Path(__file__).parent))

import httpx
from mcp.server.fastmcp import FastMCP
from dotenv import load_dotenv

load_dotenv(Path(__file__).parent / ".env", override=True)

AGENT_URL = os.environ.get("ENTERPRISE_AGENT_URL", "http://localhost:8000")
TIMEOUT = 120  # seconds — pipelines can take a while

mcp = FastMCP("Enterprise Agent", log_level="WARNING")


def _call_agent(message: str, user_id: str = "discord") -> dict:
    """POST to the enterprise agent API and return the parsed response."""
    with httpx.Client(timeout=TIMEOUT) as client:
        resp = client.post(
            f"{AGENT_URL}/v1/chat",
            json={"message": message, "user_id": user_id, "channel": "discord"},
        )
        resp.raise_for_status()
        return resp.json()


def _fmt_bi_result(data: dict) -> str:
    """Format BI pipeline result for Discord (2000 char limit)."""
    result = data.get("result", {})
    kpis = result.get("kpis", [])
    summary = result.get("exec_summary", "")
    pipeline = data.get("pipeline_used", "")
    cost = data.get("estimated_cost_usd", 0)

    lines = [f"**Enterprise BI Analysis** (pipeline: {pipeline}, cost: ${cost:.4f})"]

    if kpis:
        lines.append(f"\n**KPIs discovered ({len(kpis)}):**")
        for k in kpis[:5]:
            lines.append(f"• **{k['name']}** — {k.get('definition','')[:80]}")
            lines.append(f"  Formula: `{k.get('formula','')[:60]}`")

    dashboard = result.get("dashboard_config", {})
    if dashboard.get("charts"):
        lines.append(f"\n**Dashboard:** {dashboard.get('title','')} ({len(dashboard['charts'])} charts)")
        for c in dashboard["charts"][:3]:
            lines.append(f"• {c['type']}: {c['kpi']} by {c.get('x_axis','')}")

    if summary:
        short = textwrap.shorten(summary.replace("**", ""), width=400, placeholder="…")
        lines.append(f"\n**Exec Summary:**\n{short}")

    approval = result.get("approval_required", False)
    if approval:
        lines.append("\n⚠️ **Human approval required before release.**")

    return "\n".join(lines)[:1900]


def _fmt_sw_result(data: dict) -> str:
    """Format SW delivery pipeline result for Discord."""
    result = data.get("result", {})
    artifacts = result.get("artifacts", {})
    steps = result.get("steps", {})
    cost = data.get("estimated_cost_usd", 0)

    lines = [f"**Software Delivery Pipeline** (cost: ${cost:.4f})"]

    arch = artifacts.get("architecture", {})
    if arch.get("architecture"):
        style = arch["architecture"].get("style", "")
        stack = arch["architecture"].get("tech_stack", "")
        lines.append(f"\n**Architecture:** {style} | Stack: {stack}")

    sec = steps.get("04_security_review", {})
    if sec:
        risk = sec.get("risk_score", "?")
        lines.append(f"**Security score:** {risk}/10")

    deploy = steps.get("09_deploy", {})
    if deploy.get("deployment_plan"):
        plan = deploy["deployment_plan"]
        if isinstance(plan, list):
            lines.append(f"\n**Deployment steps ({len(plan)}):**")
            for s in plan[:4]:
                lines.append(f"• {str(s)[:80]}")

    if result.get("approval_required"):
        lines.append("\n⚠️ **Approval required before deploying to production.**")

    lines.append("\n✅ **Artifacts generated:** architecture, backend code, Dockerfile, K8s manifests, tests")
    return "\n".join(lines)[:1900]


# ─────────────────────────── MCP TOOLS ────────────────────────────────────

@mcp.tool()
def bi_analysis(request: str) -> str:
    """
    Run the 10-step BI pipeline on a business request.
    Use for: KPI discovery, dashboard generation, executive reports,
    data analysis, revenue/churn/metrics questions.

    Args:
        request: Natural language BI request
                 e.g. "Show Q1 revenue by region for executives"
    """
    try:
        data = _call_agent(request)
        return _fmt_bi_result(data)
    except httpx.ConnectError:
        return "❌ Enterprise agent is offline. Make sure it's running: `cd openclaw-enterprise-agent && python -m api.gateway`"
    except Exception as e:
        return f"❌ BI pipeline error: {e}"


@mcp.tool()
def sw_delivery(request: str) -> str:
    """
    Run the 10-step software delivery pipeline.
    Use for: building APIs, generating code, architecture design,
    creating Dockerfiles, Kubernetes manifests, test generation.

    Args:
        request: Description of the software to build
                 e.g. "Build a REST API for user authentication"
    """
    try:
        data = _call_agent(request)
        return _fmt_sw_result(data)
    except httpx.ConnectError:
        return "❌ Enterprise agent is offline."
    except Exception as e:
        return f"❌ SW delivery error: {e}"


@mcp.tool()
def data_catalog() -> str:
    """
    List available data sources and tables in the enterprise data platform.
    Use when asked about connected databases, available data, or data sources.
    """
    try:
        with httpx.Client(timeout=10) as client:
            resp = client.get(f"{AGENT_URL}/v1/data/catalog")
            resp.raise_for_status()
            catalog = resp.json()

        lines = ["**Enterprise Data Catalog**"]
        for source, info in catalog.items():
            if "error" in source:
                lines.append(f"• ⚠️ {source}: {info}")
            else:
                lines.append(f"• **{source}:** {info}")
        return "\n".join(lines)
    except httpx.ConnectError:
        return "❌ Enterprise agent is offline."
    except Exception as e:
        return f"❌ Catalog error: {e}"


@mcp.tool()
def agent_stats() -> str:
    """
    Show enterprise agent runtime statistics: LLM calls, token usage, cost, cache hit rate.
    Use when asked about agent performance, costs, or usage.
    """
    try:
        with httpx.Client(timeout=10) as client:
            resp = client.get(f"{AGENT_URL}/v1/stats")
            resp.raise_for_status()
            stats = resp.json()

        model_stats = stats.get("model_stats", {})
        lines = [
            "**Enterprise Agent Stats**",
            f"• LLM calls: {model_stats.get('total_calls', 0)}",
            f"• Input tokens: {model_stats.get('total_input_tokens', 0):,}",
            f"• Output tokens: {model_stats.get('total_output_tokens', 0):,}",
            f"• Estimated cost: ${model_stats.get('estimated_cost_usd', 0):.4f}",
            f"• Cache hit rate: {stats.get('cache_hit_rate', 0):.1%}",
        ]
        calls_by_model = model_stats.get("calls_by_model", {})
        if calls_by_model:
            lines.append("• Models used: " + ", ".join(f"{m} ({n})" for m, n in calls_by_model.items()))
        return "\n".join(lines)
    except httpx.ConnectError:
        return "❌ Enterprise agent is offline."
    except Exception as e:
        return f"❌ Stats error: {e}"


@mcp.tool()
def enterprise_health() -> str:
    """Check if the enterprise agent platform is running and healthy."""
    try:
        with httpx.Client(timeout=5) as client:
            resp = client.get(f"{AGENT_URL}/v1/health")
            resp.raise_for_status()
            return f"✅ Enterprise agent is **online** at {AGENT_URL} — version {resp.json().get('version', '?')}"
    except httpx.ConnectError:
        return f"❌ Enterprise agent is **offline** (tried {AGENT_URL})"
    except Exception as e:
        return f"⚠️ Health check error: {e}"


@mcp.tool()
def skill_hub_query(query: str) -> str:
    """
    Query the enterprise skill hub — discover which skills agents are using.
    Use when asked about skills, agent capabilities, or what tools are available.

    Args:
        query: One of:
               • "list" — all 33 installed skills
               • "role:<name>" — skills for a specific agent role (e.g. "role:backend")
               • "task:<name>" — skills for a task type (e.g. "task:debugging")
               • "swarm:<name>" — skills for a whole swarm (e.g. "swarm:qa")
               • "search:<keyword>" — search across all skills
               • "summary" — full hub overview
    """
    try:
        sys.path.insert(0, str(Path(__file__).parent))
        from skill_hub.registry import SkillRegistry
        reg = SkillRegistry()

        q = query.strip().lower()

        if q == "list":
            skills = reg.list_all()
            lines = [f"**Skill Hub — {len(skills)} installed skills**"]
            for s in skills:
                meta = reg.get(s)
                lines.append(f"• **{s}** — {meta.description[:60] if meta else ''}")
            return "\n".join(lines)[:1900]

        if q == "summary":
            info = reg.summary()
            lines = [
                f"**Skill Hub Summary**",
                f"• Total skills: {info['total_skills']}",
                f"• Roles mapped: {len(info['roles'])}",
                f"• Task types: {len(info['tasks'])}",
                f"• Swarms: {', '.join(info['swarms'])}",
                f"\n**All skills:** {', '.join(info['loaded'])}",
            ]
            return "\n".join(lines)[:1900]

        if q.startswith("role:"):
            role = q[5:].strip()
            card = reg.role_card(role)
            lines = [f"**Skills for role: {role}** ({card['skill_count']} skills)"]
            for s in card["skills"]:
                lines.append(f"• **{s['name']}** — {s['description']}")
            if not card["skills"]:
                lines.append(f"No skills mapped. Available roles: {', '.join(reg.list_roles()[:10])}")
            return "\n".join(lines)[:1900]

        if q.startswith("task:"):
            task = q[5:].strip()
            card = reg.task_card(task)
            lines = [f"**Skills for task: {task}** ({card['skill_count']} skills)"]
            for s in card["skills"]:
                lines.append(f"• **{s['name']}** — {s['description']}")
            if not card["skills"]:
                lines.append(f"No skills mapped. Available tasks: {', '.join(reg.list_tasks()[:12])}")
            return "\n".join(lines)[:1900]

        if q.startswith("swarm:"):
            swarm = q[6:].strip()
            skills = reg.get_for_swarm(swarm)
            lines = [f"**Skills for swarm: {swarm}** ({len(skills)} unique skills)"]
            for s in skills:
                lines.append(f"• **{s.name}** — {s.description[:70]}")
            if not skills:
                lines.append("No skills found for that swarm.")
            return "\n".join(lines)[:1900]

        if q.startswith("search:"):
            keyword = q[7:].strip()
            results = reg.search(keyword)
            lines = [f"**Search '{keyword}' — {len(results)} results**"]
            for s in results[:8]:
                lines.append(f"• **{s.name}** — {s.description[:70]}")
            return "\n".join(lines)[:1900]

        # Default: treat as search
        results = reg.search(q)
        if results:
            lines = [f"**Skill search: '{q}' — {len(results)} results**"]
            for s in results[:6]:
                lines.append(f"• **{s.name}** — {s.description[:70]}")
            return "\n".join(lines)[:1900]

        return f"Unknown query '{query}'. Try: list, summary, role:<name>, task:<name>, swarm:<name>, search:<keyword>"
    except Exception as e:
        return f"❌ Skill hub error: {e}"


@mcp.tool()
def duckdb_query(sql: str) -> str:
    """
    Run an analytical SQL query via DuckDB (in-process, no server needed).
    Use for: ad-hoc analytics on CSV/Parquet files, aggregations, KPI calculations.
    Skill: duckdb-query (skills.sh)

    Args:
        sql: DuckDB SQL query. Use read_csv_auto('path/file.csv') to load files inline.
             e.g. "SELECT month, SUM(revenue) FROM read_csv_auto('/data/sales.csv') GROUP BY 1"
    """
    try:
        import sys
        sys.path.insert(0, str(Path(__file__).parent))
        from data.connectors.duckdb_connector import DuckDBConnector
        conn = DuckDBConnector()
        df = conn.query(sql)
        if df.empty:
            return "No results returned."
        lines = ["**DuckDB Query Result**", f"```", df.to_string(index=False, max_rows=20), "```"]
        return "\n".join(lines)[:1900]
    except Exception as e:
        return f"❌ DuckDB error: {e}"


@mcp.tool()
def generate_xlsx_report(request: str) -> str:
    """
    Run a BI analysis and export the results as an XLSX report.
    Skill: xlsx (anthropics/skills via skills.sh)
    Use when asked to create, download, or export a spreadsheet/Excel report.

    Args:
        request: What data to include in the report, e.g. "Q1 revenue and churn report"
    """
    try:
        import io, base64, json as _json
        data = _call_agent(request)
        result = data.get("result", {})
        kpis = result.get("kpis", [])
        steps = result.get("steps", {})

        # Build XLSX in memory using openpyxl if available, else return JSON
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.Workbook()

            # Sheet 1: KPIs
            ws = wb.active
            ws.title = "KPIs"
            ws.append(["KPI Name", "Definition", "Formula"])
            ws["A1"].font = Font(bold=True)
            ws["B1"].font = Font(bold=True)
            ws["C1"].font = Font(bold=True)
            for kpi in kpis:
                ws.append([kpi.get("name", ""), kpi.get("definition", ""), kpi.get("formula", "")])

            # Sheet 2: Exec Summary
            ws2 = wb.create_sheet("Executive Summary")
            ws2.append(["Executive Summary"])
            ws2["A1"].font = Font(bold=True, size=14)
            summary = result.get("exec_summary", "")
            for i, line in enumerate(summary.split("\n")[:30], start=2):
                ws2[f"A{i}"] = line

            # Save to buffer
            buf = io.BytesIO()
            wb.save(buf)
            b64 = base64.b64encode(buf.getvalue()).decode()
            return (
                f"**XLSX Report Generated** ({len(kpis)} KPIs, {len(summary)} chars summary)\n"
                f"Base64 data (paste into file decoder):\n```\n{b64[:200]}...\n```\n"
                f"_(Full report: {len(buf.getvalue())} bytes)_"
            )
        except ImportError:
            # openpyxl not installed — return JSON summary
            report = {"kpis": kpis, "exec_summary": result.get("exec_summary", ""), "steps": list(steps.keys())}
            return f"**Report (JSON — install openpyxl for XLSX)**\n```json\n{_json.dumps(report, indent=2)[:1400]}\n```"
    except httpx.ConnectError:
        return "❌ Enterprise agent is offline."
    except Exception as e:
        return f"❌ XLSX report error: {e}"


if __name__ == "__main__":
    mcp.run(transport="stdio")
